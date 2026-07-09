import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import numpy as np
import os
import io
import re
import zlib
import random
import json
from concurrent.futures import ThreadPoolExecutor
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from generate_mock_data import ORANGE3_TYPES, ORANGE3_ROLES, grade_marketing_risk

# ==========================================
# 1. 초기 설정 및 기업 데이터 풀
# ==========================================
st.set_page_config(page_title="융합형 주식 시뮬레이터", page_icon="📈", layout="wide", initial_sidebar_state="expanded")

# ==========================================
# 🎨 테마 시스템 (다크/라이트 모드)
# ==========================================
DARK_PALETTE = {
    "bg": "#0E1117", "panel": "#171C26", "text": "#E6EAF2", "sub": "#8B95A5",
    "border": "rgba(255,255,255,0.09)",
    "sidebar": "linear-gradient(180deg, #141B2B 0%, #0E1117 100%)",
    "hero": "linear-gradient(120deg, #131A2B 0%, #1B2A4A 55%, #26335E 100%)",
    "title": "linear-gradient(90deg, #7CB8FF, #B79CFF)",
    "card": "linear-gradient(135deg, rgba(124,184,255,0.07), rgba(183,156,255,0.04))",
    "accent": "#7CB8FF", "badge_blue": "#7CB8FF", "badge_red": "#FF8A8A", "badge_purple": "#C9B4FF",
}
LIGHT_PALETTE = {
    "bg": "#F4F6FB", "panel": "#FFFFFF", "text": "#1A2233", "sub": "#5B6472",
    "border": "rgba(23,32,64,0.13)",
    "sidebar": "linear-gradient(180deg, #FFFFFF 0%, #F1F4FC 100%)",
    "hero": "linear-gradient(120deg, #E9F0FF 0%, #EEE9FF 55%, #FDEFF6 100%)",
    "title": "linear-gradient(90deg, #2563EB, #7C3AED)",
    "card": "linear-gradient(135deg, rgba(37,99,235,0.05), rgba(124,58,237,0.04))",
    "accent": "#2563EB", "badge_blue": "#2563EB", "badge_red": "#DC2626", "badge_purple": "#7C3AED",
}

IS_LIGHT_MODE = st.session_state.get("light_mode", False)
_P = LIGHT_PALETTE if IS_LIGHT_MODE else DARK_PALETTE

# 팔레트를 CSS 변수로 주입 (라이트/다크 전환 시 이 블록만 바뀜)
st.markdown(f"""
<style>
:root {{
    --bg: {_P['bg']}; --panel: {_P['panel']}; --text: {_P['text']}; --sub: {_P['sub']};
    --border: {_P['border']}; --sidebar-bg: {_P['sidebar']};
    --hero-grad: {_P['hero']}; --title-grad: {_P['title']}; --card-grad: {_P['card']};
    --accent: {_P['accent']};
    --badge-blue: {_P['badge_blue']}; --badge-red: {_P['badge_red']}; --badge-purple: {_P['badge_purple']};
}}
html {{ color-scheme: {'light' if IS_LIGHT_MODE else 'dark'}; }}
</style>
""", unsafe_allow_html=True)

# 전역 디자인 (팔레트 변수 기반 - 라이트/다크 공용)
st.markdown("""
<style>
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.min.css');
    html, body, [data-testid="stAppViewContainer"], [data-testid="stSidebar"], button, input, textarea {
        font-family: 'Pretendard', 'Noto Sans KR', -apple-system, 'Malgun Gothic', sans-serif;
    }
    /* Streamlit 아이콘 폰트 보호 - 폰트를 덮어쓰면 아이콘이 글자로 깨져 보임 */
    [data-testid="stIconMaterial"], [class*="material-symbols"] {
        font-family: 'Material Symbols Rounded' !important;
    }

    [data-testid="stAppViewContainer"] { background: var(--bg); }
    [data-testid="stHeader"] { background: transparent; }
    [data-testid="stSidebar"] { background: var(--sidebar-bg); border-right: 1px solid var(--border); }

    /* 텍스트 색상 */
    [data-testid="stMarkdownContainer"] p, [data-testid="stMarkdownContainer"] li,
    [data-testid="stMarkdownContainer"] h1, [data-testid="stMarkdownContainer"] h2,
    [data-testid="stMarkdownContainer"] h3, [data-testid="stMarkdownContainer"] h4,
    [data-testid="stMarkdownContainer"] h5, [data-testid="stMarkdownContainer"] table td,
    [data-testid="stMarkdownContainer"] table th { color: var(--text); }
    [data-testid="stMarkdownContainer"] table td, [data-testid="stMarkdownContainer"] table th {
        border-color: var(--border) !important;
    }
    [data-testid="stCaptionContainer"], [data-testid="stCaptionContainer"] p { color: var(--sub) !important; }
    hr { border-color: var(--border); }

    /* 메트릭 카드 */
    [data-testid="stMetric"] {
        background: var(--card-grad);
        border: 1px solid var(--border);
        border-radius: 14px;
        padding: 14px 18px;
    }
    [data-testid="stMetricLabel"] p { color: var(--sub); }
    [data-testid="stMetricValue"] { color: var(--text); font-weight: 800; letter-spacing: -0.5px; }
    [data-testid="stVerticalBlockBorderWrapper"] { border-radius: 16px; border-color: var(--border) !important; }

    /* 버튼 */
    .stButton > button, .stDownloadButton > button {
        border-radius: 10px;
        border: 1px solid var(--border);
        transition: all .15s ease;
    }
    .stButton > button[kind="secondary"], .stDownloadButton > button {
        background: var(--panel);
        color: var(--text);
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        border-color: var(--accent);
        color: var(--accent);
        transform: translateY(-1px);
    }

    /* 입력 위젯 */
    .stTextArea textarea, .stTextInput input {
        background: var(--panel) !important;
        color: var(--text) !important;
        border: 1px solid var(--border);
        border-radius: 10px;
    }
    /* 셀렉트박스: 기본 테마(다크) 배경이 라이트 모드에서도 검게 남는 문제 해결 */
    [data-testid="stSelectbox"] > div > div,
    [data-baseweb="select"] > div {
        background: var(--panel) !important;
        border-color: var(--border) !important;
        color: var(--text) !important;
    }
    [data-testid="stSelectbox"] input, [data-baseweb="select"] input { color: var(--text) !important; }
    [data-testid="stSelectbox"] svg, [data-baseweb="select"] svg { fill: var(--sub); }
    /* 셀렉트박스 펼침 메뉴 (portal로 별도 렌더링되므로 직접 지정) */
    [data-baseweb="popover"] > div, [data-baseweb="popover"] > div > div,
    [data-baseweb="menu"], ul[role="listbox"],
    [data-testid="stSelectboxVirtualDropdown"] {
        background: var(--panel) !important;
    }
    li[role="option"], [data-testid="stSelectboxVirtualDropdown"] li {
        background: var(--panel) !important;
        color: var(--text) !important;
    }
    li[role="option"]:hover, li[role="option"][aria-selected="true"] {
        background: color-mix(in srgb, var(--accent) 14%, var(--panel)) !important;
    }

    /* 탭 */
    button[data-baseweb="tab"] { font-weight: 700; font-size: 1rem; color: var(--sub); }
    button[data-baseweb="tab"][aria-selected="true"] { color: var(--accent); }

    /* 익스팬더 */
    [data-testid="stExpander"] details {
        background: var(--panel);
        border: 1px solid var(--border);
        border-radius: 12px;
    }
    [data-testid="stExpander"] summary {
        color: var(--text);
        background: var(--panel) !important;
        border-radius: 12px;
    }
    [data-testid="stExpander"] summary:hover { color: var(--accent); }
    [data-testid="stExpander"] summary span { color: var(--text); }

    /* 히어로 배너 */
    .hero-banner {
        background: var(--hero-grad);
        border: 1px solid var(--border);
        border-radius: 18px;
        padding: 24px 30px;
        margin-bottom: 14px;
    }
    .hero-head { display: flex; align-items: center; gap: 12px; }
    .hero-emoji { font-size: 1.8rem; }
    .hero-title {
        font-size: 1.85rem; font-weight: 800; letter-spacing: -0.5px;
        background: var(--title-grad);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    }
    .hero-badge {
        display: inline-block; padding: 5px 14px; border-radius: 999px;
        font-size: 0.9rem; font-weight: 600; margin-right: 8px; margin-top: 12px;
    }
    .badge-blue { background: color-mix(in srgb, var(--badge-blue) 15%, transparent); color: var(--badge-blue); }
    .badge-red { background: color-mix(in srgb, var(--badge-red) 15%, transparent); color: var(--badge-red); }
    .badge-purple { background: color-mix(in srgb, var(--badge-purple) 15%, transparent); color: var(--badge-purple); }

    /* 섹션 타이틀 (왼쪽 포인트 바) */
    .sec-title {
        display: flex; align-items: center; gap: 8px;
        font-weight: 800; font-size: 1.05rem; color: var(--text); margin: 6px 0 12px;
    }
    .sec-title::before {
        content: ''; width: 4px; height: 18px; border-radius: 2px;
        background: var(--title-grad);
    }

    /* Orange3 워크플로우 애니메이션 (분석 가이드) */
    .o3-canvas {
        display: flex; align-items: center; padding: 22px 18px;
        background: color-mix(in srgb, var(--accent) 5%, transparent);
        border: 1.5px dashed var(--border); border-radius: 14px;
        overflow-x: auto; margin: 8px 0 14px;
    }
    .o3-node { display: flex; flex-direction: column; align-items: center; min-width: 92px; }
    .o3-circle {
        width: 58px; height: 58px; border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        font-size: 1.6rem;
        box-shadow: 0 4px 14px rgba(0,0,0,0.18);
        animation: o3pulse 2.6s ease-in-out infinite;
    }
    .o3-label { margin-top: 8px; font-size: 0.8rem; font-weight: 700; color: var(--text); white-space: nowrap; }
    .o3-link {
        position: relative; flex: 1 1 70px; min-width: 70px; height: 0;
        border-top: 2px dashed var(--sub); margin: 0 6px; top: -14px;
    }
    .o3-dot {
        position: absolute; top: -6px; left: 0; width: 10px; height: 10px;
        border-radius: 50%; background: var(--accent);
        animation: o3flow 1.8s linear infinite;
    }
    .o3-dot.d2 { animation-delay: 0.9s; }
    @keyframes o3flow {
        0% { left: 0; opacity: 0; }
        15% { opacity: 1; }
        85% { opacity: 1; }
        100% { left: calc(100% - 10px); opacity: 0; }
    }
    @keyframes o3pulse {
        0%, 100% { transform: scale(1); }
        50% { transform: scale(1.06); }
    }

    /* 라인 차트 등장 애니메이션: 선이 왼쪽부터 그려지고 점이 뒤따라 나타남 (레이스 차트 등) */
    [data-testid="stPlotlyChart"] { animation: chartRise .5s ease-out; }
    [data-testid="stPlotlyChart"] .scatterlayer .trace path.js-line {
        stroke-dasharray: 3500;
        stroke-dashoffset: 3500;
        animation: drawLine 2s linear forwards;
    }
    [data-testid="stPlotlyChart"] .scatterlayer .trace .points path {
        opacity: 0;
        animation: fadeDot .6s ease-out forwards;
        animation-delay: 1.4s;
    }
    @keyframes chartRise { from { opacity: 0; transform: translateY(10px); } to { opacity: 1; transform: none; } }
    @keyframes drawLine { to { stroke-dashoffset: 0; } }
    @keyframes fadeDot { to { opacity: 1; } }

    /* 분기별 이슈 타임라인 (레이스 차트 아래) */
    .issue-strip { display: flex; flex-direction: column; gap: 6px; margin: 4px 0 8px; }
    .issue-item { display: flex; align-items: center; gap: 10px; font-size: 0.86rem; }
    .issue-q {
        flex: 0 0 auto;
        background: color-mix(in srgb, var(--accent) 14%, transparent);
        color: var(--accent); font-weight: 700;
        padding: 2px 10px; border-radius: 999px; font-size: 0.78rem;
    }
    .issue-t { color: var(--text); }

    /* 정보 카드 그리드 (분석 가이드 - 저학년 가독성용 시각 자료) */
    .g-grid {
        display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
        gap: 10px; margin: 6px 0 16px;
    }
    .g-card {
        background: var(--panel); border: 1px solid var(--border);
        border-radius: 14px; padding: 14px 14px 12px;
        transition: transform .15s ease;
    }
    .g-card:hover { transform: translateY(-2px); }
    .g-ico { font-size: 1.5rem; }
    .g-tit { font-weight: 800; color: var(--text); margin: 6px 0 3px; font-size: 0.95rem; }
    .g-desc { font-size: 0.82rem; color: var(--sub); line-height: 1.5; }
</style>
""", unsafe_allow_html=True)

# 교사 인증 번호: .streamlit/secrets.toml에 TEACHER_PIN = "원하는번호" 추가로 변경 가능 (기본값 2026)
try:
    TEACHER_PIN = str(st.secrets.get("TEACHER_PIN", "2026"))
except Exception:
    TEACHER_PIN = "2026"

# 산업군별 포인트 색상 (기업 카드 칩)
INDUSTRY_COLORS = {
    '반도체/전자': '#F59E0B', 'IT/소프트웨어': '#3B82F6', '바이오/제약': '#10B981',
    '엔터/미디어': '#EC4899', '자동차/모빌리티': '#8B5CF6', '금융/은행': '#14B8A6',
    '식음료/소비재': '#EF4444', '에너지/화학': '#84CC16',
}

# 상태를 영구 저장할 파일 경로 설정
STATE_FILE = "data/simulation_state.json"

company_pool = {
    '삼성전자': {'industry': '반도체/전자', 'price': 80000}, 'SK하이닉스': {'industry': '반도체/전자', 'price': 170000}, 'LG전자': {'industry': '반도체/전자', 'price': 100000}, '삼성전기': {'industry': '반도체/전자', 'price': 150000}, 'DB하이텍': {'industry': '반도체/전자', 'price': 50000},
    '네이버': {'industry': 'IT/소프트웨어', 'price': 190000}, '카카오': {'industry': 'IT/소프트웨어', 'price': 50000}, '엔씨소프트': {'industry': 'IT/소프트웨어', 'price': 200000}, '크래프톤': {'industry': 'IT/소프트웨어', 'price': 250000}, '삼성SDS': {'industry': 'IT/소프트웨어', 'price': 160000},
    '셀트리온': {'industry': '바이오/제약', 'price': 180000}, '삼성바이오로직스': {'industry': '바이오/제약', 'price': 800000}, '유한양행': {'industry': '바이오/제약', 'price': 70000}, '한미약품': {'industry': '바이오/제약', 'price': 300000}, 'SK바이오팜': {'industry': '바이오/제약', 'price': 90000},
    '하이브': {'industry': '엔터/미디어', 'price': 200000}, 'JYP Ent.': {'industry': '엔터/미디어', 'price': 70000}, '에스엠': {'industry': '엔터/미디어', 'price': 80000}, '와이지엔터테인먼트': {'industry': '엔터/미디어', 'price': 40000}, '스튜디오드래곤': {'industry': '엔터/미디어', 'price': 45000},
    '현대차': {'industry': '자동차/모빌리티', 'price': 240000}, '기아': {'industry': '자동차/모빌리티', 'price': 110000}, '현대모비스': {'industry': '자동차/모빌리티', 'price': 250000}, 'HL만도': {'industry': '자동차/모빌리티', 'price': 35000}, '한온시스템': {'industry': '자동차/모빌리티', 'price': 6000},
    'KB금융': {'industry': '금융/은행', 'price': 70000}, '신한지주': {'industry': '금융/은행', 'price': 50000}, '하나금융지주': {'industry': '금융/은행', 'price': 60000}, '우리금융지주': {'industry': '금융/은행', 'price': 15000}, '카카오뱅크': {'industry': '금융/은행', 'price': 25000},
    'CJ제일제당': {'industry': '식음료/소비재', 'price': 300000}, '아모레퍼시픽': {'industry': '식음료/소비재', 'price': 130000}, '농심': {'industry': '식음료/소비재', 'price': 400000}, '오리온': {'industry': '식음료/소비재', 'price': 90000}, 'LG생활건강': {'industry': '식음료/소비재', 'price': 350000},
    'LG에너지솔루션': {'industry': '에너지/화학', 'price': 400000}, 'SK이노베이션': {'industry': '에너지/화학', 'price': 120000}, 'S-Oil': {'industry': '에너지/화학', 'price': 75000}, '한화솔루션': {'industry': '에너지/화학', 'price': 30000}, '포스코퓨처엠': {'industry': '에너지/화학', 'price': 300000}
}

market_events = [
    {"이슈": "특별한 이슈 없음 (시장 안정세)", "impact": {}},
    {"이슈": "글로벌 인플레이션 심화 (소비 위축 및 금리 인상 기조)", "impact": {"식음료/소비재": 0.85, "엔터/미디어": 0.80, "금융/은행": 1.15}}, 
    {"이슈": "AI 혁명 가속화 (IT/반도체 슈퍼 호재)", "impact": {"반도체/전자": 1.30, "IT/소프트웨어": 1.20, "에너지/화학": 0.95}},
    {"이슈": "새로운 글로벌 전염병 변이 발생 (헬스케어 수혜)", "impact": {"바이오/제약": 1.35, "엔터/미디어": 0.75, "자동차/모빌리티": 0.80, "식음료/소비재": 1.10}},
    {"이슈": "친환경 ESG 글로벌 규제 강화 (탄소 배출 제한)", "impact": {"에너지/화학": 0.80, "자동차/모빌리티": 1.20, "반도체/전자": 0.95}},
    {"이슈": "K-콘텐츠 글로벌 대흥행 및 수출 호조", "impact": {"엔터/미디어": 1.30, "식음료/소비재": 1.15, "IT/소프트웨어": 1.05}},
    {"이슈": "국제 원자재 및 유가 폭등 (공급망 충격)", "impact": {"자동차/모빌리티": 0.75, "반도체/전자": 0.80, "에너지/화학": 1.25}},
    {"이슈": "글로벌 반도체 공급 과잉 우려 (재고 증가)", "impact": {"반도체/전자": 0.75, "자동차/모빌리티": 1.10, "IT/소프트웨어": 0.95}},
    {"이슈": "저출산 및 고령화 심화 (시니어 산업 부상)", "impact": {"바이오/제약": 1.20, "식음료/소비재": 0.85, "엔터/미디어": 0.90}},
    {"이슈": "정부 주도 기업 밸류업(주주환원) 정책 발표", "impact": {"금융/은행": 1.25, "자동차/모빌리티": 1.15, "반도체/전자": 1.05}},
    {"이슈": "중국 내수 경기 침체 및 애국소비(궈차오) 열풍", "impact": {"식음료/소비재": 0.75, "엔터/미디어": 0.80, "바이오/제약": 0.90}},
    {"이슈": "차세대 모빌리티(자율주행/UAM) 상용화 규제 철폐", "impact": {"자동차/모빌리티": 1.30, "IT/소프트웨어": 1.15, "에너지/화학": 1.10}},
    {"이슈": "가상화폐(비트코인) 역대 최고가 경신 및 핀테크 호황", "impact": {"IT/소프트웨어": 1.20, "금융/은행": 1.10, "엔터/미디어": 0.95}},
    {"이슈": "MZ세대 중심의 헬스디깅(건강관리) 트렌드 확산", "impact": {"바이오/제약": 1.15, "식음료/소비재": 1.10, "IT/소프트웨어": 1.05}},
    {"이슈": "주요국 무역 갈등 심화 (관세 폭탄 우려)", "impact": {"반도체/전자": 0.85, "자동차/모빌리티": 0.85, "에너지/화학": 0.90}}
]

# ==========================================
# 🌟 [신규] 파일 저장 및 로드 백업 시스템 함수
# ==========================================
def load_saved_file_state():
    """새로고침을 방어하기 위해 파일에서 데이터를 읽어옴"""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def get_state_dict():
    """현재 진행 상태 스냅샷 (파일 저장과 교사 화면의 백업 다운로드가 공유)"""
    return {
        "setup_complete": st.session_state.setup_complete,
        "current_quarter": st.session_state.current_quarter,
        "market_issue": st.session_state.market_issue,
        "students_companies": st.session_state.students_companies,
        "price_history": st.session_state.price_history,
        "ai_evaluations": st.session_state.ai_evaluations,
        "quarter_records": st.session_state.quarter_records,
        "generated_reports": st.session_state.generated_reports,
        "portfolios": st.session_state.portfolios
    }

def save_current_state_to_file():
    """턴이 변경되거나 세팅될 때 실시간 파일 동기화 저장"""
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(get_state_dict(), f, ensure_ascii=False, indent=4)

# 백업본 먼저 로드하기
saved_data = load_saved_file_state()

# ==========================================
# 2. 세션 상태 (Session State) 초기화 (백업본 연동)
# ==========================================
if 'setup_complete' not in st.session_state:
    st.session_state.setup_complete = saved_data.get("setup_complete", False)
if 'current_quarter' not in st.session_state:
    st.session_state.current_quarter = saved_data.get("current_quarter", 1)
if 'market_issue' not in st.session_state:
    st.session_state.market_issue = saved_data.get("market_issue", market_events[0])
if 'students_companies' not in st.session_state:
    st.session_state.students_companies = saved_data.get("students_companies", [])
if 'price_history' not in st.session_state:
    st.session_state.price_history = saved_data.get("price_history", {comp: [info['price']] for comp, info in company_pool.items()})
if 'ai_evaluations' not in st.session_state:
    st.session_state.ai_evaluations = saved_data.get("ai_evaluations", {})
if 'quarter_records' not in st.session_state:
    # 분기별 진행 기록: 시장 이슈, 조별 전략/AI 평가를 누적 저장 ('시뮬레이션 결과 CSV'의 원천 데이터)
    st.session_state.quarter_records = saved_data.get("quarter_records", [])
if 'generated_reports' not in st.session_state:
    # 마케팅 성과가 반영되어 매 분기 새로 생성되는 '월별 실적 보고서' (기업별 월 단위 행 누적)
    st.session_state.generated_reports = saved_data.get("generated_reports", {})
if 'portfolios' not in st.session_state:
    # 조별 투자 포트폴리오: 시드머니(현금) + 보유 종목(수량/평균 매수단가) + 거래 내역
    st.session_state.portfolios = saved_data.get("portfolios", {})

# ==========================================
# 3. 핵심 기능 함수 모음
# ==========================================
@st.cache_data
def load_company_df(company):
    """기업 CSV를 읽어 순수 데이터만 반환 (Orange3 자료형/역할 헤더 2줄은 자동 건너뜀)"""
    file_path = f"data/{company}_초기분석데이터.csv"
    if not os.path.exists(file_path):
        return None
    df = pd.read_csv(file_path)
    if len(df) > 0 and str(df.iloc[0, 0]).strip() in ("string", "discrete", "continuous", "time"):
        df = df.iloc[2:].reset_index(drop=True)
    return df

@st.cache_data
def load_all_companies_df():
    """40개 기업의 과거 3개년 데이터를 하나로 합친 480행 데이터 (의사결정나무 학습용).
    한 기업의 12행만으로는 나무가 우연 상관(예: ESG 평판)에 과적합되므로,
    위기 등급 규칙을 제대로 역산하려면 이 통합 데이터로 학습해야 한다."""
    frames = []
    for comp in company_pool:
        df = load_company_df(comp)
        if df is None:
            continue
        df = df.copy()
        df.insert(0, '기업명', comp)
        frames.append(df)
    return pd.concat(frames, ignore_index=True)

def _coerce_number(value):
    """CSV에서 문자열로 읽힌 숫자를 엑셀에 숫자 셀로 넣기 위한 변환"""
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return value

# 여러 기업을 합친 시트에서는 글자 열의 종류가 16개를 넘어 Orange3 Tree의
# exhaustive binarization 한도에 걸리므로, 해당 열을 meta(참고 정보)로 강등한다.
TEXT_META_OVERRIDES = {'주요 타겟층': 'meta', '주력 판매 제품': 'meta', '주요 판매 채널': 'meta'}

def _add_orange3_sheet(wb, sheet_name, df, role_overrides=None):
    """워크북에 Orange3 3줄 헤더(컬럼명/자료형/역할)가 붙은 시트를 추가하고 서식 적용"""
    roles = dict(ORANGE3_ROLES)
    if role_overrides:
        roles.update(role_overrides)
    ws = wb.create_sheet(sheet_name)
    cols = list(df.columns)
    ws.append(cols)
    ws.append([ORANGE3_TYPES.get(c, "string") for c in cols])
    ws.append([roles.get(c, "") for c in cols])
    for row in df.itertuples(index=False):
        ws.append([_coerce_number(v) for v in row])

    header_fill = PatternFill("solid", fgColor="1F2A44")
    for cell in ws[1]:
        cell.font = XLFont(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for meta_row in ws.iter_rows(min_row=2, max_row=3):
        for cell in meta_row:
            cell.font = XLFont(color="8B95A5", size=9, italic=True)
    ws.freeze_panes = "A4"
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(c)) * 1.8, 11), 30)

def build_excel(sheets):
    """(시트이름, 데이터프레임[, 역할 오버라이드]) 목록을 하나의 xlsx 파일(bytes)로 묶음"""
    wb = Workbook()
    wb.remove(wb.active)
    for item in sheets:
        _add_orange3_sheet(wb, item[0], item[1], item[2] if len(item) > 2 else None)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def clamp(value, low, high):
    return max(low, min(high, value))

def summarize_quarter_months(monthly_rows, quarter_no):
    """월별 실적 3행을 분기 요약 1행으로 집계 (매출/이익은 합계, 비율 지표는 평균)"""
    year = 2026 + (quarter_no - 1) // 4
    q = (quarter_no - 1) % 4 + 1
    mean = lambda k: round(sum(float(r[k]) for r in monthly_rows) / len(monthly_rows), 1)
    inventory, churn = mean('재고율(%)'), mean('고객 이탈률(%)')
    roas, satisfaction = mean('ROAS(%)'), mean('고객 만족도(점)')
    return {
        'Year_Quarter': f"{year}_{q}Q",
        '매출액(억원)': sum(int(r['매출액(억원)']) for r in monthly_rows),
        '영업이익(억원)': sum(int(r['영업이익(억원)']) for r in monthly_rows),
        '시장점유율(%)': mean('시장점유율(%)'),
        '주요 타겟층': monthly_rows[-1]['주요 타겟층'],
        '주력 판매 제품': monthly_rows[-1]['주력 판매 제품'],
        '주요 판매 채널': monthly_rows[-1]['주요 판매 채널'],
        '재고율(%)': inventory,
        'ROAS(%)': int(roas),
        '고객 이탈률(%)': churn,
        '해외매출비중(%)': int(mean('해외매출비중(%)')),
        '고객 만족도(점)': int(satisfaction),
        'ESG 평판(점)': int(mean('ESG 평판(점)')),
        '마케팅_위기_등급': grade_marketing_risk(inventory, churn, roas, satisfaction),
    }

def get_latest_financials(company):
    """AI 평가와 다음 분기 실적 생성의 기준이 되는 '가장 최근' 재무 상태.
    진행된 분기가 있으면 마지막 생성 실적의 분기 요약, 없으면 과거 데이터의 마지막 분기(2025_4Q)."""
    reports = st.session_state.generated_reports.get(company, [])
    if len(reports) >= 3:
        done_quarters = len(reports) // 3
        return summarize_quarter_months(reports[(done_quarters - 1) * 3: done_quarters * 3], done_quarters)
    df = load_company_df(company)
    if df is None:
        return None
    return df.iloc[-1].to_dict()

def generate_quarter_report(company, quarter_no, multiplier):
    """마케팅 평가 결과(multiplier)를 반영해 이번 분기의 '월별 실적 보고서' 3행을 생성.
    직전 분기 실적을 기준으로 지표가 연속적으로 변하며, clamp로 비현실적인 폭주를 방지.
    좋은 전략(배수>1)은 매출·점유율·ROAS·만족도를 올리고 재고율·이탈률을 낮춘다."""
    base = get_latest_financials(company)
    year = 2026 + (quarter_no - 1) // 4
    q = (quarter_no - 1) % 4 + 1
    effect = multiplier - 1  # 마케팅 성과 (-0.25 ~ +0.25 수준)

    sales_q = max(int(float(base['매출액(억원)']) * (1 + effect * 0.8 + random.uniform(-0.04, 0.04))), 10)
    prev_sales = float(base['매출액(억원)'])
    margin = float(base['영업이익(억원)']) / prev_sales if prev_sales > 0 else 0.1
    margin = clamp(margin + effect * 0.05 + random.uniform(-0.01, 0.01), 0.02, 0.35)
    share = clamp(float(base['시장점유율(%)']) + effect * 12 + random.uniform(-1.0, 1.0), 1.0, 60.0)
    inventory = clamp(float(base['재고율(%)']) - effect * 15 + random.uniform(-2.0, 2.0), 3.0, 50.0)
    roas = clamp(float(base['ROAS(%)']) + effect * 120 + random.uniform(-10, 10), 50, 400)
    churn = clamp(float(base['고객 이탈률(%)']) - effect * 8 + random.uniform(-1.0, 1.0), 1.0, 30.0)
    satisfaction = clamp(float(base['고객 만족도(점)']) + effect * 15 + random.uniform(-2, 2), 50, 100)
    overseas = clamp(float(base['해외매출비중(%)']) + random.uniform(-3, 3), 5, 90)
    esg = clamp(float(base['ESG 평판(점)']) + effect * 5 + random.uniform(-2, 2), 30, 100)

    # 분기 매출을 월별로 배분 (가중치 정규화로 합계 보존)
    weights = [random.uniform(0.8, 1.2) for _ in range(3)]
    total_weight = sum(weights)

    rows = []
    for month_idx in range(3):
        month = (q - 1) * 3 + month_idx + 1
        month_sales = int(sales_q * weights[month_idx] / total_weight)
        m_inventory = round(clamp(inventory + random.uniform(-1.5, 1.5), 3.0, 50.0), 1)
        m_roas = int(clamp(roas + random.uniform(-8, 8), 50, 400))
        m_churn = round(clamp(churn + random.uniform(-0.8, 0.8), 1.0, 30.0), 1)
        m_satisfaction = int(clamp(satisfaction + random.uniform(-1.5, 1.5), 50, 100))
        rows.append({
            'Year_Month': f"{year}_{month:02d}",
            '매출액(억원)': month_sales,
            '영업이익(억원)': int(month_sales * margin),
            '시장점유율(%)': round(clamp(share + random.uniform(-0.4, 0.4), 1.0, 60.0), 1),
            '주요 타겟층': base['주요 타겟층'],
            '주력 판매 제품': base['주력 판매 제품'],
            '주요 판매 채널': base['주요 판매 채널'],
            '재고율(%)': m_inventory,
            'ROAS(%)': m_roas,
            '고객 이탈률(%)': m_churn,
            '해외매출비중(%)': int(overseas),
            '고객 만족도(점)': m_satisfaction,
            'ESG 평판(점)': int(esg),
            '마케팅_위기_등급': grade_marketing_risk(m_inventory, m_churn, m_roas, m_satisfaction),
        })
    return rows

def get_market_rows():
    """40개 기업 전체의 '가장 최근' 재무 상태 + 현재 주가/등락률 통합 데이터 (군집화 분석용).
    학생 조 기업은 마케팅이 반영된 최신 생성 실적, 배경 기업은 과거 마지막 실적(2025_4Q) 기준."""
    rows = []
    for comp, info in company_pool.items():
        fin = get_latest_financials(comp)
        if fin is None:
            continue
        fin = dict(fin)
        fin.pop('Year_Quarter', None)
        fin.pop('Year_Month', None)
        hist = st.session_state.price_history[comp]
        prev = hist[-2] if len(hist) > 1 else hist[-1]
        pct = round((hist[-1] - prev) / prev * 100, 2) if prev > 0 else 0.0
        rows.append({
            '기업명': comp, '산업군': info['industry'],
            '현재주가(원)': hist[-1], '주가등락률(%)': pct,
            **fin
        })
    return rows

# ==========================================
# 💰 조별 주식 투자 (시드머니 · 매수/매도 · 손익 평가)
# ==========================================
SEED_MONEY_PER_STUDENT = 1_000_000    # 1인당 100만원
TEAM_SIZE = 3                          # 한 조 3명
TEAM_SEED_MONEY = SEED_MONEY_PER_STUDENT * TEAM_SIZE  # 조당 300만원

def get_portfolio(team_no):
    """조 번호의 포트폴리오를 가져오고, 처음이면 시드머니로 초기화"""
    return st.session_state.portfolios.setdefault(str(team_no), {
        "cash": TEAM_SEED_MONEY, "holdings": {}, "tx": []
    })

def portfolio_summary(team_no):
    """현금 + 보유 주식의 현재가 평가액으로 조의 총자산·손익을 계산"""
    pf = get_portfolio(team_no)
    rows, stock_value = [], 0
    for comp, h in pf["holdings"].items():
        cur = st.session_state.price_history[comp][-1]
        value = h["shares"] * cur
        cost = h["shares"] * h["avg_price"]
        stock_value += value
        rows.append({
            "종목": comp,
            "보유 수량": h["shares"],
            "매수 단가(원)": round(h["avg_price"]),
            "현재가(원)": cur,
            "평가액(원)": value,
            "평가손익(원)": round(value - cost),
            "수익률(%)": round((cur - h["avg_price"]) / h["avg_price"] * 100, 2) if h["avg_price"] > 0 else 0.0,
        })
    total = pf["cash"] + stock_value
    profit = total - TEAM_SEED_MONEY
    return {
        "cash": pf["cash"], "stock_value": round(stock_value), "total": round(total),
        "profit": round(profit), "return_pct": profit / TEAM_SEED_MONEY * 100,
        "rows": rows,
    }

def get_simulation_result_rows(company):
    """진행이 끝난 분기별로 '생성된 실적(분기 요약) + 시뮬레이션 결과'를 한 행으로 결합.
    각 행 = 마케팅이 반영된 분기 실적 + 분기 말 주가, 등락률, AI 평가배수, 시장 이슈, 제출한 전략.
    학생들이 '전략과 시장 상황이 실적과 주가에 어떤 영향을 줬는지'를 Orange3로 분석하는 용도."""
    records = st.session_state.quarter_records
    reports = st.session_state.generated_reports.get(company, [])
    if not records or not reports:
        return []
    hist = st.session_state.price_history[company]
    rows = []
    for i, rec in enumerate(records):
        monthly_rows = reports[i * 3: (i + 1) * 3]
        if len(monthly_rows) < 3 or i + 1 >= len(hist):
            break
        summary = summarize_quarter_months(monthly_rows, rec.get("quarter", i + 1))
        price_before, price_after = hist[i], hist[i + 1]
        pct = round((price_after - price_before) / price_before * 100, 2) if price_before > 0 else 0.0
        result = rec["results"].get(company, {})
        rows.append({
            **summary,
            '주가(원)': price_after,
            '주가등락률(%)': pct,
            'AI평가배수': result.get("multiplier", ""),
            '시장이슈': rec["issue"],
            '마케팅전략': str(result.get("strategy", "")).replace("\n", " ").strip(),
            'AI코멘트': str(result.get("reason", "")).replace("\n", " ").strip(),
        })
    return rows

@st.cache_data(show_spinner=False)
def build_company_excel(company, quarter):
    """조별 기업의 모든 분석 데이터를 시트별로 담은 통합 엑셀 파일 생성.
    시트: ①과거 3개년 데이터 ②월별 실적 보고서 ③시뮬레이션 결과 ④전체 시장
    ⑤전체기업 3개년(480행, 의사결정나무 학습용) — 진행 상황에 따라 3~5개
    quarter는 캐시 키 용도: 데이터는 분기가 진행될 때만 바뀌므로, 그 외의 화면 갱신에서는
    엑셀을 다시 만들지 않고 재사용한다 (매 클릭마다 통짜 재생성되던 병목 제거)."""
    sheets = [("과거3개년데이터", load_company_df(company))]
    reports = st.session_state.generated_reports.get(company, [])
    if reports:
        sheets.append(("월별실적보고서", pd.DataFrame(reports)))
    sim_rows = get_simulation_result_rows(company)
    if sim_rows:
        sheets.append(("시뮬레이션결과", pd.DataFrame(sim_rows)))
    sheets.append(("전체시장", pd.DataFrame(get_market_rows()), TEXT_META_OVERRIDES))
    sheets.append(("전체기업3개년", load_all_companies_df(), TEXT_META_OVERRIDES))
    return build_excel(sheets)

@st.cache_data(show_spinner=False)
def build_market_excel(quarter):
    """전체 시장 통합 엑셀 (quarter는 캐시 키 - 분기 진행 시에만 재생성)"""
    return build_excel([("전체시장", pd.DataFrame(get_market_rows()), TEXT_META_OVERRIDES)])

def get_financial_text_for_ai(company):
    """AI 평가에 넣을 최신 재무 상태 텍스트 (진행된 분기가 있으면 마케팅이 반영된 최신 실적 기준)"""
    fin_dict = get_latest_financials(company)
    if fin_dict is None:
        return "재무 데이터 없음"
    skip_cols = ("Year_Quarter", "Year_Month", "마케팅_위기_등급")
    return ", ".join(f"{k}: {v}" for k, v in fin_dict.items() if k not in skip_cols)

def evaluate_strategy_with_ai(company, industry, strategy, market_event, financial_data):
    if not strategy.strip():
        return 1.0, ""
    try:
        genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
        # flash 모델: pro 대비 응답이 빠르고 과금이 저렴해 수업용 반복 호출에 적합
        model = genai.GenerativeModel('gemini-2.5-flash')
        prompt = f"""
        너는 주식 시장 시뮬레이터의 냉철한 평가 AI야. 
        
        [현재 기업 및 시장 정보]
        - 기업명: {company} (산업군: {industry})
        - 시장 상황: {market_event}
        - 현재 재무 및 마케팅 지표: {financial_data}
        
        [학생들의 마케팅 전략]
        - 전략 내용: {strategy}
        
        [평가 미션]
        위의 정보를 바탕으로 학생이 제시한 전략이 얼마나 효과적인지 논리적으로 분석해라.
        - 최악의 전략: 0.75 ~ 0.95
        - 무난한 전략: 0.96 ~ 1.04
        - 훌륭한 전략: 1.05 ~ 1.25
        
        [출력 형식]
        반드시 아래 형식으로만 출력해. (인사말이나 다른 설명 절대 금지)
        첫째 줄: 주가에 곱할 '등락률' 숫자 하나 (예: 1.08)
        둘째 줄부터 아래 세 항목을 각 줄에 하나씩 작성해. 학생(청소년)이 이해하기 쉬운 말로,
        반드시 위에 제시된 실제 재무 지표 수치나 시장 상황을 근거로 들어 구체적으로 써라.
        👍 긍정 요인: (전략의 강점을 지표·시장 상황과 연결해서 1~2문장)
        ⚠️ 위험 요인: (전략의 약점이나 놓친 부분을 1~2문장)
        💡 다음 분기 조언: (다음 전략에 반영할 구체적인 팁 1문장)
        """
        response = model.generate_content(prompt)
        lines = [ln.strip() for ln in response.text.strip().split('\n') if ln.strip()]

        # 첫 줄에 인사말이나 마크다운 기호가 섞여도 숫자만 안전하게 추출
        num_match = re.search(r'\d+\.?\d*', lines[0]) if lines else None
        if num_match is None:
            raise ValueError(f"AI 응답에서 등락률 숫자를 찾지 못함: {response.text[:100]}")
        multiplier = min(max(float(num_match.group()), 0.5), 1.5)
        # 둘째 줄부터는 모두 평가 내용 (긍정/위험/조언 항목별 여러 줄)
        reason = "\n".join(lines[1:]) if len(lines) > 1 else "시장 상황 및 전략이 주가 변동에 일부 반영되었습니다."

        return multiplier, reason
    except Exception as e:
        return random.uniform(0.95, 1.05), "AI 서버 응답 지연으로 인한 시장 평균 변동치 임의 적용"

def generate_pro_stock_chart(company_name, price_history, ai_multiplier, num_points=60):
    """시뮬레이션 시작(2026년 1분기)부터 현재까지 진행된 모든 분기를 이어붙인 캔들스틱 차트.
    분기마다 crc32 고정 시드를 사용하므로 분기가 지나가도 이미 본 구간의 캔들 모양은 그대로 유지된다.
    (과거 3개년(2023~2025)은 주가 데이터가 없는 재무 데이터라 주가 차트에는 포함되지 않는다.)"""
    num_segments = max(1, len(price_history) - 1)
    all_dates, all_opens, all_highs, all_lows, all_closes = [], [], [], [], []
    for seg in range(num_segments):
        seg_year = 2026 + seg // 4
        seg_q = seg % 4 + 1
        start_price = price_history[seg]
        end_price = price_history[seg + 1] if seg + 1 < len(price_history) else price_history[seg]

        # 파이썬 내장 hash()는 실행할 때마다 값이 달라지므로(해시 랜덤화),
        # 서버를 재시작해도 차트 모양이 유지되도록 crc32로 고정 시드 생성
        fixed_seed = zlib.crc32(f"{company_name}_{seg_year}_{seg_q}".encode("utf-8"))
        np.random.seed(fixed_seed)

        trend = np.linspace(start_price, end_price, num_points)
        volatility = start_price * 0.015
        noise = np.random.normal(0, volatility, num_points)

        closes = trend + np.cumsum(noise) - np.sum(noise) * np.linspace(0, 1, num_points)
        closes[0], closes[-1] = start_price, end_price

        opens = np.empty_like(closes)
        opens[0] = start_price
        opens[1:] = closes[:-1] + np.random.normal(0, volatility*0.2, num_points-1)

        highs = np.maximum(opens, closes) + np.abs(np.random.normal(0, volatility*0.4, num_points))
        lows = np.minimum(opens, closes) - np.abs(np.random.normal(0, volatility*0.4, num_points))

        start_month = (seg_q - 1) * 3 + 1
        all_dates.append(pd.bdate_range(start=f"{seg_year}-{start_month:02d}-01", periods=num_points))
        all_opens.append(opens)
        all_highs.append(highs)
        all_lows.append(lows)
        all_closes.append(closes)

    dates = all_dates[0].append(all_dates[1:]) if len(all_dates) > 1 else all_dates[0]
    opens, highs = np.concatenate(all_opens), np.concatenate(all_highs)
    lows, closes = np.concatenate(all_lows), np.concatenate(all_closes)

    fig = go.Figure(data=[go.Candlestick(
        x=dates,
        open=opens, high=highs, low=lows, close=closes,
        increasing_line_color='#FF4B4B', decreasing_line_color='#3D8BFF',
        increasing_fillcolor='#FF4B4B', decreasing_fillcolor='#3D8BFF',
        name="주가 흐름",
        hovertemplate="<b>%{x|%Y년 %m월 %d일}</b><br><br>" + 
                      "시가(시작): %{open:,.0f}원<br>" +
                      "고가(최고): %{high:,.0f}원<br>" +
                      "저가(최저): %{low:,.0f}원<br>" +
                      "종가(마감): %{close:,.0f}원<extra></extra>" 
    )])
    
    if ai_multiplier != 1.0:
        percent_str = f"+{round((ai_multiplier-1)*100, 1)}%" if ai_multiplier > 1 else f"{round((ai_multiplier-1)*100, 1)}%"
        bg_color = '#FF4B4B' if ai_multiplier > 1 else '#3D8BFF'
        fig.add_annotation(
            x=dates[-1], y=price_history[-1],
            text=f"<b>AI Score: {percent_str}</b>",
            showarrow=True, arrowhead=2, arrowsize=1.5, arrowwidth=2, arrowcolor=bg_color,
            ax=-50, ay=-40,
            font=dict(size=12, color="white"),
            bgcolor=bg_color, bordercolor=bg_color, borderpad=4, borderwidth=1, opacity=0.9
        )
    
    # 라이트/다크 모드에 맞춘 차트 색상
    light = st.session_state.get("light_mode", False)
    grid_color = 'rgba(0,0,0,0.08)' if light else 'rgba(255,255,255,0.07)'
    tick_color = '#6B7280' if light else '#8B95A5'
    hover_bg, hover_fg = ('#FFFFFF', '#1A2233') if light else ('#1A2233', '#E6EAF2')

    fig.update_layout(
        height=280, margin=dict(l=0, r=40, t=10, b=10),
        xaxis=dict(
            showgrid=False, visible=True, type='date', tickformat="%Y년 %m월",
            tickfont=dict(color=tick_color, size=11), fixedrange=False, rangeslider=dict(visible=False)
        ),
        yaxis=dict(
            showgrid=True, gridcolor=grid_color, side='right',
            tickfont=dict(color=tick_color, size=11),
            tickformat=",.0f", fixedrange=False
        ),
        dragmode="zoom",
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        hovermode='x unified',
        hoverlabel=dict(bgcolor=hover_bg, bordercolor=grid_color, font=dict(color=hover_fg))
    )
    return fig

TEAM_LINE_COLORS = ['#7CB8FF', '#FF8A8A', '#69DB7C', '#C9B4FF']

def generate_price_history_chart(companies, normalize=False, height=300):
    """시뮬레이션 시작부터 현재까지의 분기별 주가 흐름 라인 차트.
    각 분기 지점에 해당 분기의 시장 이슈가 툴팁으로 표시된다.
    normalize=True면 시작가를 100으로 환산 → 조별 수익률 레이스(경쟁) 차트로 사용."""
    light = st.session_state.get("light_mode", False)
    grid_color = 'rgba(0,0,0,0.08)' if light else 'rgba(255,255,255,0.07)'
    tick_color = '#6B7280' if light else '#8B95A5'
    hover_bg, hover_fg = ('#FFFFFF', '#1A2233') if light else ('#1A2233', '#E6EAF2')
    records = st.session_state.quarter_records

    fig = go.Figure()
    for idx, comp in enumerate(companies):
        hist = st.session_state.price_history[comp]
        labels, hovers = ['시작'], ['시뮬레이션 시작가']
        for i in range(1, len(hist)):
            year, q = 2026 + (i - 1) // 4, (i - 1) % 4 + 1
            labels.append(f"{year}년 {q}분기")
            issue = records[i - 1]['issue'] if i - 1 < len(records) else ''
            hovers.append(f"🚨 {issue}")
        if normalize:
            # 시작 주가 = 100 기준 지수. 호버에 '지수 (수익률 %)'를 함께 보여줘 단위 혼동을 줄인다
            y = [round(p / hist[0] * 100, 2) for p in hist]
            point_text = [f"{v - 100:+.1f}%" for v in y]
            value_fmt = '%{y:.1f} (수익률 %{text})'
        else:
            y = hist
            point_text = None
            value_fmt = '%{y:,.0f}원'
        # 통합 툴팁에서 같은 분기 이슈가 회사 수만큼 반복되지 않도록 첫 번째 선에만 이슈를 붙인다
        if idx == 0:
            hovertemplate = '<b>' + comp + '</b> ' + value_fmt + '<br>%{customdata}<extra></extra>'
        else:
            hovertemplate = '<b>' + comp + '</b> ' + value_fmt + '<extra></extra>'
        fig.add_trace(go.Scatter(
            x=labels, y=y, mode='lines+markers', name=comp,
            line=dict(width=3, color=TEAM_LINE_COLORS[idx % len(TEAM_LINE_COLORS)]),
            marker=dict(size=8),
            text=point_text,
            customdata=hovers,
            hovertemplate=hovertemplate
        ))
    if normalize:
        fig.add_hline(y=100, line_dash='dot', line_color=grid_color)

    fig.update_layout(
        height=height, margin=dict(l=0, r=20, t=10, b=10),
        showlegend=len(companies) > 1,
        legend=dict(orientation='h', y=1.15, font=dict(color=tick_color)),
        xaxis=dict(showgrid=False, tickfont=dict(color=tick_color, size=11)),
        yaxis=dict(showgrid=True, gridcolor=grid_color, tickfont=dict(color=tick_color, size=11),
                   tickformat=',.0f', side='right',
                   title=dict(text='수익률 지수 (시작=100)' if normalize else '주가(원)',
                              font=dict(color=tick_color, size=11))),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', hovermode='x unified',
        hoverlabel=dict(bgcolor=hover_bg, font=dict(color=hover_fg))
    )
    return fig

def generate_industry_impact_chart(changes):
    """이번 분기 산업군별 평균 등락률 가로 막대 차트.
    '이번 시장 이슈가 어떤 산업을 웃고 울렸는지'를 한눈에 보여준다 (군집화 수업 연계)."""
    light = st.session_state.get("light_mode", False)
    grid_color = 'rgba(0,0,0,0.08)' if light else 'rgba(255,255,255,0.07)'
    tick_color = '#6B7280' if light else '#8B95A5'

    sums, counts = {}, {}
    for comp, pct in changes.items():
        ind = company_pool[comp]['industry']
        sums[ind] = sums.get(ind, 0) + pct
        counts[ind] = counts.get(ind, 0) + 1
    industries = sorted(sums, key=lambda k: sums[k] / counts[k])
    avgs = [round(sums[i] / counts[i], 2) for i in industries]
    colors = [INDUSTRY_COLORS.get(i, '#7CB8FF') for i in industries]

    fig = go.Figure(go.Bar(
        x=avgs, y=industries, orientation='h',
        marker_color=colors,
        texttemplate='%{x:+.2f}%', textposition='outside', cliponaxis=False,
        textfont=dict(color=tick_color),
        hovertemplate='%{y}: 평균 %{x:+.2f}%<extra></extra>'
    ))
    fig.update_layout(
        height=310, margin=dict(l=0, r=55, t=10, b=10),
        xaxis=dict(showgrid=True, gridcolor=grid_color, zerolinecolor=grid_color,
                   tickfont=dict(color=tick_color, size=11), ticksuffix='%'),
        yaxis=dict(tickfont=dict(color=tick_color, size=12)),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)'
    )
    return fig

def orange3_flow(nodes):
    """Orange3 캔버스처럼 위젯 노드가 연결되고 데이터가 점으로 흘러가는 CSS 애니메이션 HTML 생성.
    nodes: (이모지, 위젯이름, 배경색) 튜플 리스트"""
    parts = []
    for i, (icon, label, color) in enumerate(nodes):
        parts.append(
            f'<div class="o3-node"><div class="o3-circle" style="background:{color};">{icon}</div>'
            f'<div class="o3-label">{label}</div></div>'
        )
        if i < len(nodes) - 1:
            parts.append('<div class="o3-link"><span class="o3-dot"></span><span class="o3-dot d2"></span></div>')
    return '<div class="o3-canvas">' + ''.join(parts) + '</div>'

def company_card_header(company, industry):
    """기업 카드 상단: 기업명 + 산업군 컬러 칩"""
    chip = INDUSTRY_COLORS.get(industry, '#7CB8FF')
    return (
        f"<div style='display:flex; align-items:center; gap:10px; margin-bottom:6px;'>"
        f"<span style='font-size:1.15rem; font-weight:800; color:var(--text);'>{company}</span>"
        f"<span style='background:{chip}26; color:{chip}; padding:3px 10px; border-radius:999px;"
        f" font-size:0.78rem; font-weight:700;'>{industry}</span></div>"
    )

def info_cards(cards):
    """아이콘+제목+짧은 설명 카드 그리드 HTML. 긴 줄글 대신 시각 자료로 전달하기 위한 헬퍼.
    cards: (이모지, 제목, 짧은설명(HTML 가능), 포인트색) 튜플 리스트"""
    items = ''.join(
        f'<div class="g-card" style="border-top:3px solid {color};">'
        f'<div class="g-ico">{icon}</div><div class="g-tit">{title}</div>'
        f'<div class="g-desc">{desc}</div></div>'
        for icon, title, desc, color in cards
    )
    return f'<div class="g-grid">{items}</div>'

# ==========================================
# 4. 앱 화면 구성 (UI)
# ==========================================
st.sidebar.markdown("""
<div style="padding: 4px 0 12px 0;">
    <div style="display:flex; align-items:center; gap:8px;">
        <span style="font-size:1.3rem;">📈</span>
        <span style="font-size:1.25rem; font-weight:800; letter-spacing:-0.5px;
                    background: var(--title-grad);
                    -webkit-background-clip:text; -webkit-text-fill-color:transparent;">융합 주식 시뮬레이터</span>
    </div>
    <div style="font-size:0.82rem; color:var(--sub); margin-top:4px;">데이터 분석 × 마케팅 전략 수업</div>
</div>
""", unsafe_allow_html=True)

st.sidebar.toggle("🌞 라이트 모드", key="light_mode", help="화면 분위기를 밝게/어둡게 전환합니다.")
st.sidebar.divider()

# 현재 진행 상황을 어느 화면에서든 볼 수 있도록 사이드바에 상시 표시
if st.session_state.setup_complete:
    _sb_year = 2026 + (st.session_state.current_quarter - 1) // 4
    _sb_q = (st.session_state.current_quarter - 1) % 4 + 1
    st.sidebar.caption(f"🗓️ 현재 **{_sb_year}년 {_sb_q}분기** 진행 중 · 완료된 분기 {len(st.session_state.quarter_records)}개")

view_mode = st.sidebar.radio("모드 선택", ["👨‍🎓 학생 화면 (대시보드)", "⚙️ 교사 화면 (전략 적용)", "📚 분석 가이드 (Orange3 사용법)"])

st.sidebar.divider()
# 교사 전용 기능: 인증된 경우에만 초기화 버튼 노출 (학생 오조작 방지)
if st.session_state.get("teacher_auth", False):
    st.sidebar.caption("🔓 교사 모드 인증됨")
    if st.sidebar.button("🚨 전체 데이터 초기화", use_container_width=True):
        # 리셋 버튼을 눌렀을 때에만 백업 파일을 강제 삭제
        if os.path.exists(STATE_FILE):
            os.remove(STATE_FILE)
        st.cache_data.clear()  # 캐시된 엑셀/CSV도 함께 비워 이전 게임 데이터가 남지 않도록
        st.session_state.clear()
        st.rerun()
    if st.sidebar.button("🔒 교사 모드 잠그기", use_container_width=True):
        st.session_state.teacher_auth = False
        st.rerun()
else:
    st.sidebar.caption("🔒 교사 기능(전략 평가·데이터 초기화)은 교사 화면에서 인증 후 사용할 수 있습니다.")

# ------------------------------------------
# [화면 D] 분석 가이드 (세팅 여부와 관계없이 항상 열람 가능)
# ------------------------------------------
if view_mode == "📚 분석 가이드 (Orange3 사용법)":
    st.markdown("""
    <div class="hero-banner">
        <div class="hero-head"><span class="hero-emoji">📚</span><span class="hero-title">데이터 분석 가이드</span></div>
        <span class="hero-badge badge-blue">Orange3 사용법</span>
        <span class="hero-badge badge-purple">모델 · 평가 · 군집화</span>
        <span class="hero-badge badge-red">전략 처방전</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div class='sec-title'>🎮 게임은 이렇게 돌아가요</div>", unsafe_allow_html=True)
    st.markdown(info_cards([
        ("📥", "1. 데이터 받기", "학생 화면에서<br>통합 Excel 다운로드", "#FFA94D"),
        ("🍊", "2. Orange3 분석", "위기 신호와<br>기회 찾기", "#74C0FC"),
        ("✍️", "3. 전략 쓰기", "데이터를 근거로<br>마케팅 기획", "#69DB7C"),
        ("📈", "4. 결과 확인", "주가와 실적으로<br>검증하고 반복!", "#F783AC"),
    ]), unsafe_allow_html=True)

    st.markdown("<div class='sec-title'>📦 받는 데이터 한눈에 보기</div>", unsafe_allow_html=True)
    st.markdown(info_cards([
        ("📜", "과거3개년데이터", "2023~2025 성적표<br>첫 전략의 근거", "#FFA94D"),
        ("🗓️", "월별실적보고서", "내 마케팅 성과가<br>매달 기록됨", "#74C0FC"),
        ("🎯", "시뮬레이션결과", "전략 → 주가 변화<br>복기용", "#B197FC"),
        ("🌐", "전체시장", "40개 기업 비교<br>그룹 나누기용", "#69DB7C"),
        ("🌳", "전체기업3개년", "40개 기업 × 3년<br>의사결정나무용", "#38D9A9"),
    ]), unsafe_allow_html=True)
    st.caption("⚠️ 엑셀의 2~3번째 줄은 Orange3 설정값이에요. 지우면 안 돼요!")

    st.markdown("<div class='sec-title'>🍊 Orange3의 위젯 서랍 5개</div>", unsafe_allow_html=True)
    st.markdown(info_cards([
        ("📁", "Data", "데이터 불러오기<br>표로 보기", "#FFA94D"),
        ("📊", "Visualize", "그래프로<br>눈으로 확인", "#FFD43B"),
        ("🌳", "Model", "컴퓨터가<br>규칙을 학습", "#69DB7C"),
        ("✅", "Evaluate", "모델 실력<br>채점하기", "#74C0FC"),
        ("🧩", "Unsupervised", "비슷한 것끼리<br>그룹 만들기", "#B197FC"),
    ]), unsafe_allow_html=True)
    st.caption("사용법은 딱 하나! 왼쪽 서랍에서 위젯을 끌어다 놓고 → 선으로 잇는다.")

    st.divider()

    with st.expander("📥 STEP 1. 데이터 불러오기 (Data)", expanded=True):
        st.markdown(orange3_flow([
            ("📁", "File", "#FFA94D"), ("📋", "Data Table", "#74C0FC"), ("💡", "분석 시작!", "#FFD43B"),
        ]), unsafe_allow_html=True)
        st.markdown(info_cards([
            ("①", "File 놓기", "Data 서랍에서<br>캔버스로 드래그", "#FFA94D"),
            ("②", "파일 선택", "더블클릭 → <b>...</b> 버튼<br>→ 우리 Excel 선택", "#FFA94D"),
            ("③", "시트 고르기", "드롭다운에서<br>분석할 시트 선택", "#FFA94D"),
            ("④", "표 확인", "Data Table 연결<br>→ 눈으로 확인", "#FFA94D"),
        ]), unsafe_allow_html=True)
        st.markdown("""
        🖱️ **그대로 따라 하기**
        1. Orange3를 켜고 **New**를 눌러 빈 캔버스를 열어요.
        2. 왼쪽 **Data** 서랍에서 **File** 위젯을 캔버스 가운데로 끌어다 놓아요.
        3. File을 **더블클릭** → **📂(...) 버튼** → 다운받은 `우리회사_통합분석.xlsx`를 선택해요.
        4. 파일 이름 아래 **시트 드롭다운**에서 `과거3개년데이터`를 골라요.
        5. 가운데 표에서 `마케팅_위기_등급` 줄을 찾아요. Role 칸이 **target**이 아니면 그 칸을 **더블클릭**해 target으로 바꾸고 아래 **Apply**를 눌러요.
        6. **Data Table**을 끌어다 놓고, File의 **오른쪽 반원(⌒)** 을 잡아끌어 Data Table의 **왼쪽 반원**에 놓으면 선이 이어져요.
        7. Data Table을 더블클릭했을 때 **12줄(12개 분기)** 이 보이면 성공! ✅

        🎬 **실제 조작 영상**: [Orange 공식 유튜브](https://www.youtube.com/@OrangeDataMining/playlists) · [위젯 설명서](https://orangedatamining.com/widget-catalog/) · 유튜브 검색 **"오렌지3 사용법"**
        """)

    with st.expander("📊 STEP 2. 눈으로 관찰하기 (Visualize)"):
        st.markdown(orange3_flow([
            ("📁", "File", "#FFA94D"), ("📊", "Distributions", "#FFD43B"), ("📦", "Box Plot", "#FFC078"),
        ]), unsafe_allow_html=True)
        st.markdown(info_cards([
            ("📊", "Distributions", "지표가 어떻게<br>퍼져 있는지", "#FFD43B"),
            ("📦", "Box Plot", "등급별로 지표가<br>얼마나 다른지", "#FFC078"),
            ("✨", "Scatter Plot", "두 지표 사이의<br>관계 보기", "#F783AC"),
        ]), unsafe_allow_html=True)
        st.markdown("""
        🖱️ **그대로 따라 하기**: File → **Box Plot** 선 잇기 → Box Plot 더블클릭 → **Variable**에서 `재고율(%)` 선택 → **Subgroups**를 `마케팅_위기_등급`으로!
        등급(안전/주의/위험)마다 상자의 위치가 **확 다르게** 보이는 지표가 있다면, 그게 위기를 가르는 범인 후보예요. Variable을 하나씩 바꿔 가며 비교해 보세요.

        💡 **미션**: 우리 회사의 **가장 아픈 지표**를 찾아라! (재고율? 이탈률? ROAS?)
        """)

    with st.expander("🌳 STEP 3. 규칙 배우기 (Model) - 의사결정나무"):
        st.markdown(orange3_flow([
            ("📁", "File (전체기업3개년)", "#FFA94D"), ("🌳", "Tree", "#69DB7C"), ("🔍", "Tree Viewer", "#38D9A9"),
        ]), unsafe_allow_html=True)
        st.markdown(info_cards([
            ("①", "새 File 꺼내기", "시트는 꼭<br><b>전체기업3개년</b>!", "#69DB7C"),
            ("②", "Tree 연결", "File → Tree →<br>Tree Viewer 잇기", "#69DB7C"),
            ("③", "갈림길 읽기", "예: 이탈률 > 15<br>→ 위기 가지로", "#69DB7C"),
        ]), unsafe_allow_html=True)
        st.markdown("""
        ⚠️ **여기서는 새 File 위젯에 `전체기업3개년` 시트를 열어요!**
        우리 회사 12줄만으로 배우면 컴퓨터가 우연에 속아 엉뚱한 갈림길(ESG, 매출액…)을 만들어요.
        40개 기업 × 3년 = 480줄로 배워야 진짜 규칙을 찾아냅니다.

        🖱️ **그대로 따라 하기**
        1. **File 위젯을 새로 하나** 끌어다 놓고, 같은 통합 Excel을 선택 → 시트를 `전체기업3개년`으로!
        2. **Model** 서랍에서 **Tree**를 끌어와 새 File과 선으로 이어요.
        3. **Visualize** 서랍의 **Tree Viewer**를 끌어와 Tree와 이어요.
        4. Tree Viewer를 더블클릭 → 나무 그림 등장! 맨 위 갈림길부터 조건을 읽어 내려가요.
        5. 읽는 법: 갈림길에 `고객 이탈률(%) > 15` 라고 써 있으면 → "이탈률이 15를 넘는 회사는 이쪽 가지로 간다"는 뜻이에요.
        6. 우리 회사의 **최신 분기(2025_4Q) 숫자**를 들고 갈림길을 직접 따라가 보세요. 어느 잎(안전/주의/위험)에 도착하나요?

        - 나무의 **갈림길 조건** = 위기를 가르는 기준! STEP 2의 위험 신호 기준(30/15/120/70)과 비교해 보세요.
        - 🌲 **Random Forest**: 나무 여러 그루의 투표 → 더 정확해요 (대신 나무 그림은 없음)
        - 👥 **kNN**: "너랑 비슷한 회사들이 위험하면, 너도 위험!"
        """)

    with st.expander("✅ STEP 4. 모델 채점하기 (Evaluate)"):
        st.markdown(orange3_flow([
            ("📁", "File", "#FFA94D"), ("🎓", "Test & Score", "#74C0FC"), ("🔢", "Confusion Matrix", "#5C9EFF"),
        ]), unsafe_allow_html=True)
        st.markdown(info_cards([
            ("🎓", "Test & Score", "모델 성적표<br><b>CA = 정확도</b> (1.0 만점)", "#74C0FC"),
            ("🔢", "Confusion Matrix", "뭘 맞히고<br>뭘 헷갈렸는지 표", "#5C9EFF"),
            ("🔮", "Predictions", "새 데이터의 등급을<br>모델이 예측", "#B197FC"),
        ]), unsafe_allow_html=True)
        st.markdown("""
        🖱️ **그대로 따라 하기**: **Evaluate** 서랍에서 **Test & Score**를 꺼낸 뒤, **File(전체기업3개년)과 Tree를 둘 다** Test & Score에 이어요.
        (데이터 선 1개 + 모델 선 1개 = **총 2개의 선**이 들어가야 해요!) 더블클릭하면 **CA(정확도)** 점수가 보여요. 1.0이 만점!

        - 480줄로 배우면 **0.9 근처**의 높은 점수가 나와요. File을 우리 회사(12줄)로 바꿔 보면? 뚝 떨어져요 → **데이터가 많을수록 컴퓨터는 똑똑해져요!**
        - Tree와 kNN을 같이 연결하면 **어떤 모델이 더 똑똑한지 대결**도 가능!
        - **Confusion Matrix**를 Test & Score 뒤에 이으면 "위험을 안전으로 착각한 횟수"까지 표로 보여요.
        """)

    with st.expander("🧩 STEP 5. 그룹 나누기 (Unsupervised) - K-평균 군집화"):
        st.markdown(orange3_flow([
            ("🌐", "File (전체시장)", "#69DB7C"), ("🧩", "k-Means", "#B197FC"), ("✨", "Scatter Plot", "#F783AC"),
        ]), unsafe_allow_html=True)
        st.markdown(info_cards([
            ("①", "전체시장 시트", "40개 기업 데이터<br>불러오기", "#B197FC"),
            ("②", "k-Means", "그룹 수(k)는<br>3~5부터 시작", "#B197FC"),
            ("③", "Scatter Plot", "<b>Color = Cluster</b>로!<br>그룹이 색깔로 보임", "#B197FC"),
        ]), unsafe_allow_html=True)
        st.markdown("""
        🖱️ **그대로 따라 하기**: File 위젯의 시트를 `전체시장`으로 바꾸고 → **Unsupervised** 서랍의 **k-Means**를 이어요 → **Number of clusters**를 3으로 → **Scatter Plot**을 이어요 → Scatter Plot에서 **Color를 Cluster로!**
        40개 기업이 색깔별 그룹으로 나뉘어 보이면 성공이에요.

        - 💡 **질문**: 우리 회사는 어느 그룹? 잘나가는 그룹과 **무엇이 다른가**?
        - 🌿 **Hierarchical Clustering**: **Distances** 위젯을 먼저 연결하면 가계도처럼 묶어줘요.
        - 군집화는 정답(target)이 필요 없어요. 컴퓨터가 스스로 비슷한 회사를 찾아냅니다.
        """)

    with st.expander("🆘 막힐 때 보는 문제 해결 모음 (Q&A)"):
        st.markdown("""
        | 이런 증상이라면 | 이렇게 해결해요 |
        | --- | --- |
        | 위젯끼리 선이 안 그어져요 | 위젯 **오른쪽 끝의 반원(⌒)** 을 클릭한 채 끌어서, 다음 위젯의 **왼쪽 반원**에 놓아요. 위젯 몸통을 잡으면 이동만 돼요. |
        | 시트 선택 칸이 안 보여요 | File 위젯에서 **xlsx 파일을 먼저** 선택하면 그 아래에 시트 드롭다운이 나타나요. |
        | `마케팅_위기_등급`에 target 표시가 없어요 | File 위젯 가운데 표에서 그 줄의 **Role 칸을 더블클릭** → **target**으로 바꾸고 → **Apply**를 눌러요. |
        | Tree Viewer가 텅 비어 있어요 | File → Tree → Tree Viewer로 **선이 모두 이어졌는지**, target이 지정됐는지 확인해요. |
        | Test & Score에 점수가 안 나와요 | File(데이터)과 Tree(모델)를 **둘 다** Test & Score에 이어야 해요. 선이 2개인지 세어 보세요. |
        | 표의 줄 수가 이상하게 적어요 | 엑셀에서 2~3번째 줄(자료형/역할)을 지웠다면, 시뮬레이터에서 파일을 **다시 다운로드**하세요. |
        | 실수로 위젯을 지웠어요 | **Ctrl+Z**로 되돌리거나, 같은 위젯을 다시 끌어다 놓고 선만 이으면 돼요. |
        """)

    st.divider()

    st.markdown("<div class='sec-title'>💊 지표별 전략 처방전</div>", unsafe_allow_html=True)
    st.markdown("""
    | 발견한 문제 | 전략 방향 |
    | --- | --- |
    | 재고율 높음 | 재고 소진 프로모션 · 한정 특가 |
    | 이탈률 높음 | 멤버십 · 재구매 쿠폰 · AS 강화 |
    | ROAS 낮음 | 광고 채널 바꾸기 · 비효율 광고 중단 |
    | 만족도 낮음 | 품질 개선 · 고객 의견 반영 |
    | 점유율 낮음 | 새 타겟층 공략 · 차별화 |
    | 해외비중 낮음 | 수출 전략 (시장 이슈 먼저 확인!) |
    """)

    st.markdown("<div class='sec-title'>🤖 AI 평가 잘 받는 법 4가지</div>", unsafe_allow_html=True)
    st.markdown(info_cards([
        ("🔢", "숫자 근거 대기", "\"재고율 34%라서<br>→ 할인 프로모션\"", "#FFA94D"),
        ("📰", "이슈와 연결", "지금 핫이슈가<br>우리에게 호재? 악재?", "#74C0FC"),
        ("🎯", "타겟 정확히", "데이터에 나온<br>주요 타겟층 겨냥", "#69DB7C"),
        ("🏢", "우리 회사답게", "주력 제품·채널과<br>어울리는 전략", "#F783AC"),
    ]), unsafe_allow_html=True)

# ------------------------------------------
# [화면 A] 초기 세팅
# ------------------------------------------
elif not st.session_state.setup_complete:
    st.markdown("""
    <div class="hero-banner">
        <div class="hero-head"><span class="hero-emoji">⚙️</span><span class="hero-title">시뮬레이터 조별 기업 세팅</span></div>
        <span class="hero-badge badge-blue">STEP 1. 각 조가 운영할 기업을 선택하세요</span>
    </div>
    """, unsafe_allow_html=True)
    st.info("각 조가 담당할 4개의 기업을 선택해 주세요. (서로 다른 기업이어야 합니다)")
    
    all_companies = list(company_pool.keys())
    col1, col2 = st.columns(2)
    with col1:
        team1 = st.selectbox("1조", all_companies, index=0)
        team2 = st.selectbox("2조", all_companies, index=1)
    with col2:
        team3 = st.selectbox("3조", all_companies, index=2)
        team4 = st.selectbox("4조", all_companies, index=3)
        
    if st.button("✅ 시뮬레이션 시작"):
        selected = [team1, team2, team3, team4]
        if len(set(selected)) < 4:
            st.error("⚠️ 같은 기업을 두 조 이상이 선택했습니다. 조별로 서로 다른 기업을 골라 주세요.")
        else:
            st.session_state.students_companies = selected
            st.session_state.setup_complete = True
            save_current_state_to_file() # 🌟 파일에 강제 저장
            st.rerun()

# ------------------------------------------
# [화면 B] 학생 대시보드
# ------------------------------------------
elif view_mode == "👨‍🎓 학생 화면 (대시보드)":
    display_year = 2026 + (st.session_state.current_quarter - 1) // 4
    display_q = (st.session_state.current_quarter - 1) % 4 + 1

    current_issue_text = st.session_state.market_issue["이슈"] if isinstance(st.session_state.market_issue, dict) else st.session_state.market_issue
    st.markdown(f"""
    <div class="hero-banner">
        <div class="hero-head"><span class="hero-emoji">📈</span><span class="hero-title">글로벌 마켓 시뮬레이션</span></div>
        <span class="hero-badge badge-blue">🗓️ {display_year}년 {display_q}분기</span>
        <span class="hero-badge badge-red">🚨 {current_issue_text}</span>
    </div>
    """, unsafe_allow_html=True)
    
    with st.expander("📊 캔들스틱(봉차트) 읽는 법 & 조작 팁 - 클릭해서 확인하세요!"):
        st.markdown("""
        * 🔴 **빨간색 캔들 (양봉):** 당일 시초가 대비 주가가 올랐을 때 나타납니다.
        * 🔵 **파란색 캔들 (음봉):** 당일 시초가 대비 주가가 내렸을 때 나타납니다.
        * 🔍 **확대해서 보기 (Zoom):** 마우스로 차트 위를 **클릭 & 드래그** 하거나, **마우스 휠을 굴려보세요!** 돋보기처럼 확대되어 구체적인 '일차(Day)'별 흐름을 볼 수 있습니다. 원래 크기로 돌아오려면 그래프 위에서 **더블클릭** 하세요.
        * 🕰️ **차트 범위:** 시뮬레이션 시작(2026년 1월)부터 지금까지 진행된 **모든 분기가 이어져** 보입니다. 그 이전(2023~2025)은 주가 대신 **과거 재무 데이터**(통합 Excel)로 분석해요.
        """)

    if st.session_state.quarter_records:
        with st.expander("📜 지금까지의 시장 이슈 히스토리 - 거시 경제 흐름을 복기해 보세요!"):
            for rec in st.session_state.quarter_records:
                st.markdown(f"- **{rec['label']}**: {rec['issue']}")

    st.divider()
    
    # st.tabs는 화면이 갱신될 때마다 첫 탭으로 되돌아가서(기업 선택 시 우리 조 화면으로 튕김),
    # 선택 상태가 유지되는 라디오 버튼으로 화면을 전환한다
    student_view = st.radio(
        "보기 선택",
        ["📊 우리 조 집중 분석", "🌐 전체 시장 동향 (배경 기업 탐색)"],
        horizontal=True, key="student_view", label_visibility="collapsed"
    )

    if student_view == "📊 우리 조 집중 분석":
        # 💰 조별 투자 성과: 시드머니(조당 3,000만원)로 산 주식의 총자산·손익 현황
        _team_count = len(st.session_state.students_companies)
        if any(get_portfolio(no)["tx"] for no in range(1, _team_count + 1)):
            st.markdown(
                f"<div class='sec-title'>💰 조별 투자 성과 — 시드머니 {TEAM_SEED_MONEY:,}원, 얼마나 불렸을까?</div>",
                unsafe_allow_html=True
            )
            _invest = [
                (team_no, comp, portfolio_summary(team_no))
                for team_no, comp in enumerate(st.session_state.students_companies, start=1)
            ]
            _ranked = sorted(_invest, key=lambda x: x[2]["total"], reverse=True)
            _medals = ["🥇", "🥈", "🥉", "🏅"]
            _inv_cols = st.columns(len(_ranked))
            for _rank, (team_no, comp, s) in enumerate(_ranked):
                with _inv_cols[_rank]:
                    st.metric(
                        label=f"{_medals[_rank]} {team_no}조 총자산",
                        value=f"{s['total']:,}원",
                        delta=f"{s['profit']:+,}원 ({s['return_pct']:+.2f}%)"
                    )
            with st.expander("📂 조별 보유 종목 상세 — 어떤 주식을 얼마에 사서, 지금 얼마가 됐나?"):
                for team_no, comp, s in _invest:
                    st.markdown(
                        f"**{team_no}조 · {comp}** — 💵 현금 {s['cash']:,}원 · 📦 주식 {s['stock_value']:,}원 · "
                        f"🏦 총자산 {s['total']:,}원 ({s['return_pct']:+.2f}%)"
                    )
                    if s["rows"]:
                        st.dataframe(pd.DataFrame(s["rows"]), use_container_width=True, hide_index=True)
                    else:
                        st.caption("보유 종목 없음 (전액 현금)")
            st.divider()

        # 조별 누적 수익률 리더보드 (시작가 대비 현재가)
        if st.session_state.current_quarter > 1:
            st.markdown("<div class='sec-title'>🏆 조별 누적 수익률 순위</div>", unsafe_allow_html=True)
            perf = []
            for team_no, comp in enumerate(st.session_state.students_companies, start=1):
                hist = st.session_state.price_history[comp]
                ret = (hist[-1] - hist[0]) / hist[0] * 100 if hist[0] > 0 else 0
                perf.append((team_no, comp, ret))
            perf.sort(key=lambda x: x[2], reverse=True)
            medals = ["🥇", "🥈", "🥉", "🏅"]
            rank_cols = st.columns(4)
            for rank, (team_no, comp, ret) in enumerate(perf):
                with rank_cols[rank]:
                    st.metric(label=f"{medals[rank]} {team_no}조 · {comp}", value=f"{ret:+.2f}%")

            st.markdown("<div class='sec-title'>🏁 조별 수익률 레이스 — 시작 주가를 100으로 환산한 지수</div>", unsafe_allow_html=True)
            st.caption("💡 단위 읽는 법: 시작 주가를 100이라고 놓았을 때의 크기예요. **110 = +10% 수익, 90 = -10% 손실!** 주가가 다른 기업끼리도 공평하게 비교할 수 있어요.")
            st.plotly_chart(
                generate_price_history_chart(st.session_state.students_companies, normalize=True, height=320),
                use_container_width=True, config={'displayModeBar': False}
            )
            # 분기별 시장 이슈 타임라인: 호버 없이도 항상 보이도록 차트 아래에 배지로 표시
            if st.session_state.quarter_records:
                issue_items = ''.join(
                    f"<div class='issue-item'><span class='issue-q'>{rec['label']}</span>"
                    f"<span class='issue-t'>🚨 {rec['issue']}</span></div>"
                    for rec in st.session_state.quarter_records
                )
                st.markdown(f"<div class='issue-strip'>{issue_items}</div>", unsafe_allow_html=True)
            st.caption("어느 이슈에서 순위가 뒤집혔나요? 그래프의 분기 점에 마우스를 올리면 지수와 함께 정확한 수익률(%)도 볼 수 있어요.")
            st.divider()

        cols = st.columns(2)
        for idx, company in enumerate(st.session_state.students_companies):
            prev_price = st.session_state.price_history[company][-2] if len(st.session_state.price_history[company]) > 1 else st.session_state.price_history[company][-1]
            current_price = st.session_state.price_history[company][-1]
            price_diff = current_price - prev_price
            percent_diff = (price_diff / prev_price) * 100 if prev_price > 0 else 0
            
            ai_data = st.session_state.ai_evaluations.get(company, {"score": 1.0, "reason": ""})
            ai_score = ai_data["score"]
            ai_reason = ai_data["reason"]
            
            with cols[idx % 2]:
                with st.container(border=True):
                    st.markdown(company_card_header(company, company_pool[company]['industry']), unsafe_allow_html=True)
                    
                    st.metric(
                        label="현재 주가",
                        value=f"{int(current_price):,}원",
                        delta=f"{int(price_diff):,}원 ({percent_diff:+.2f}%)"
                    )
                    
                    if st.session_state.current_quarter > 1 and ai_reason:
                        # 항목별(긍정/위험/조언) 여러 줄 코멘트를 줄바꿈해서 표시
                        reason_md = ai_reason.replace("\n", "  \n")
                        if ai_score >= 1.0:
                            st.success(f"📈 **AI 평가 리포트**  \n{reason_md}")
                        else:
                            st.error(f"📉 **AI 평가 리포트**  \n{reason_md}")
                    
                    chart = generate_pro_stock_chart(company, st.session_state.price_history[company], ai_score)
                    st.plotly_chart(chart, use_container_width=True, config={'scrollZoom': True, 'displayModeBar': False})

                    if len(st.session_state.price_history[company]) > 1:
                        with st.expander("📈 전체 기간 주가 흐름 보기 (시작~현재)"):
                            st.plotly_chart(
                                generate_price_history_chart([company], height=260),
                                use_container_width=True, config={'displayModeBar': False}
                            )

                    st.download_button(
                        label="📊 통합 분석 파일 (Excel) - 과거데이터·월별실적·시뮬레이션결과·전체시장·전체기업3개년",
                        data=build_company_excel(company, st.session_state.current_quarter),
                        file_name=f"{company}_통합분석_{display_year}Q{display_q}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_excel_{company}_{st.session_state.current_quarter}",
                        use_container_width=True
                    )
                    
    else:
        st.markdown("<div class='sec-title'>🔍 거시 경제에 따른 다른 기업들의 흐름을 분석해 보세요</div>", unsafe_allow_html=True)

        st.download_button(
            label="📊 전체 40개 기업 통합 데이터 (Excel, 최신 실적+현재 주가) - Orange3 군집화 분석용",
            data=build_market_excel(st.session_state.current_quarter),
            file_name=f"전체시장_{display_year}Q{display_q}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_market_{st.session_state.current_quarter}",
            use_container_width=True
        )

        changes = {}
        for c in company_pool.keys():
            hist = st.session_state.price_history[c]
            p_prev = hist[-2] if len(hist) > 1 else hist[-1]
            p_curr = hist[-1]
            pct = (p_curr - p_prev) / p_prev * 100 if p_prev > 0 else 0
            changes[c] = pct
            
        sorted_changes = sorted(changes.items(), key=lambda x: x[1], reverse=True)
        top_3 = sorted_changes[:3]
        bottom_3 = sorted_changes[-3:]
        bottom_3.reverse()
        
        col_up, col_down = st.columns(2)
        with col_up:
            st.success("🔥 **이번 분기 급등주 Top 3**")
            for rank, (c_name, pct) in enumerate(top_3):
                st.markdown(f"{rank+1}. **{c_name}** (+{pct:.2f}%)")
        with col_down:
            st.error("❄️ **이번 분기 급락주 Top 3**")
            for rank, (c_name, pct) in enumerate(bottom_3):
                st.markdown(f"{rank+1}. **{c_name}** ({pct:.2f}%)")

        if st.session_state.current_quarter > 1:
            st.markdown("<div class='sec-title'>🏭 산업군별 평균 등락률 - 이번 이슈는 어떤 산업을 움직였을까?</div>", unsafe_allow_html=True)
            st.plotly_chart(
                generate_industry_impact_chart(changes),
                use_container_width=True, config={'displayModeBar': False}
            )
            st.caption("시장 이슈와 산업군의 관계를 찾아보세요. Orange3 군집화 분석과 연결하면 더 재미있어요!")

        st.divider()
        
        background_companies = [c for c in company_pool.keys() if c not in st.session_state.students_companies]
        selected_bg = st.selectbox("📊 차트를 확인할 기업을 선택하세요 (우리 조 제외 전체 기업)", background_companies)
        
        if selected_bg:
            bg_prev = st.session_state.price_history[selected_bg][-2] if len(st.session_state.price_history[selected_bg]) > 1 else st.session_state.price_history[selected_bg][-1]
            bg_curr = st.session_state.price_history[selected_bg][-1]
            bg_diff = bg_curr - bg_prev
            bg_pct = (bg_diff / bg_prev) * 100 if bg_prev > 0 else 0
            
            with st.container(border=True):
                st.markdown(company_card_header(selected_bg, company_pool[selected_bg]['industry']), unsafe_allow_html=True)
                st.metric(
                    label="현재 주가",
                    value=f"{int(bg_curr):,}원",
                    delta=f"{int(bg_diff):,}원 ({bg_pct:+.2f}%)"
                )
                
                bg_chart = generate_pro_stock_chart(selected_bg, st.session_state.price_history[selected_bg], 1.0)
                st.plotly_chart(bg_chart, use_container_width=True, config={'scrollZoom': True, 'displayModeBar': False})

# ------------------------------------------
# [화면 C] 교사 화면 (전략 적용)
# ------------------------------------------
elif view_mode == "⚙️ 교사 화면 (전략 적용)":
    # 🔒 교사 인증 게이트: 학생이 전략 평가/분기 진행을 조작하지 못하도록 PIN 잠금
    if not st.session_state.get("teacher_auth", False):
        st.markdown("""
        <div class="hero-banner">
            <div class="hero-head"><span class="hero-emoji">🔒</span><span class="hero-title">교사 전용 화면</span></div>
            <span class="hero-badge badge-red">인증이 필요합니다</span>
        </div>
        """, unsafe_allow_html=True)
        st.info("전략 평가와 분기 진행은 수업 진행자(교사)만 할 수 있습니다. 인증 번호를 입력해 주세요.")
        pin_input = st.text_input("교사 인증 번호 (PIN)", type="password", max_chars=12)
        if st.button("🔓 인증하기", type="primary"):
            if pin_input == TEACHER_PIN:
                st.session_state.teacher_auth = True
                st.rerun()
            else:
                st.error("인증 번호가 올바르지 않습니다.")
        st.stop()

    display_year = 2026 + (st.session_state.current_quarter - 1) // 4
    display_q = (st.session_state.current_quarter - 1) % 4 + 1

    st.markdown(f"""
    <div class="hero-banner">
        <div class="hero-head"><span class="hero-emoji">🛠️</span><span class="hero-title">분기별 전략 평가 시스템</span></div>
        <span class="hero-badge badge-blue">🗓️ {display_year}년 {display_q}분기 기획안 평가</span>
        <span class="hero-badge badge-purple">🤖 AI가 시장 상황·최신 실적과 함께 분석합니다</span>
    </div>
    """, unsafe_allow_html=True)

    # 클라우드 배포 시 서버 재시작으로 진행 상태가 사라질 수 있으므로 백업/복원 수단을 제공
    with st.expander("☁️ 진행 상태 백업/복원 (클라우드 배포 시 필수)"):
        st.caption(
            "클라우드(Streamlit Cloud 등)에 배포하면 서버가 재시작될 때 진행 상태가 사라질 수 있습니다. "
            "**분기를 진행할 때마다 백업 파일을 내려받아 두고**, 다음 수업에서 게임이 초기화되어 있으면 그 파일로 복원하세요."
        )
        st.download_button(
            "💾 진행 상태 백업 다운로드 (.json)",
            json.dumps(get_state_dict(), ensure_ascii=False, indent=2).encode("utf-8"),
            file_name=f"시뮬레이션_백업_{display_year}년_{display_q}분기.json", mime="application/json",
            key=f"dl_state_{st.session_state.current_quarter}", use_container_width=True
        )
        uploaded_state = st.file_uploader("복원할 백업 파일(.json) 선택", type=["json"], key="state_restore_file")
        if uploaded_state is not None:
            if st.button("♻️ 이 백업으로 복원하기 (현재 진행 상태를 덮어씁니다)", use_container_width=True):
                try:
                    restored = json.load(uploaded_state)
                except (json.JSONDecodeError, UnicodeDecodeError):
                    st.error("JSON 파일을 읽을 수 없습니다. 이 화면에서 내려받은 백업 파일이 맞는지 확인하세요.")
                else:
                    required = {"setup_complete", "current_quarter", "students_companies", "price_history"}
                    missing = required - set(restored)
                    if missing:
                        st.error("백업 파일 형식이 아닙니다. 누락된 항목: " + ", ".join(sorted(missing)))
                    else:
                        with open(STATE_FILE, "w", encoding="utf-8") as f:
                            json.dump(restored, f, ensure_ascii=False, indent=4)
                        st.cache_data.clear()
                        st.session_state.clear()
                        st.session_state.teacher_auth = True  # 복원 직후 PIN 재입력 없이 이어서 진행
                        st.rerun()

    st.divider()

    # 교사 기능을 두 개의 작업 공간으로 분리 (마케팅 평가 ↔ 주식 거래 혼동 방지)
    # st.tabs는 rerun마다 첫 탭으로 튕기므로, 선택이 유지되는 라디오로 전환한다
    teacher_view = st.radio(
        "작업 선택",
        ["📝 마케팅 전략 평가 (분기 진행)", "💰 주식 거래 (시드머니)"],
        horizontal=True, key="teacher_view", label_visibility="collapsed"
    )

    # ==========================================
    # [교사 탭 1] 💰 주식 거래 — 이 탭을 그리면 아래 마케팅 화면은 건너뛴다(st.stop)
    # ==========================================
    if teacher_view == "💰 주식 거래 (시드머니)":
        st.markdown(
            f"<div class='sec-title'>💰 조별 주식 거래 — 시드머니 조당 {TEAM_SEED_MONEY:,}원 (1인 {SEED_MONEY_PER_STUDENT:,}원 × {TEAM_SIZE}명)</div>",
            unsafe_allow_html=True
        )
        if "trade_msg" in st.session_state:
            st.success(st.session_state.pop("trade_msg"))

        team_labels = [f"{no}조 · {comp}" for no, comp in enumerate(st.session_state.students_companies, start=1)]
        trade_c1, trade_c2, trade_c3 = st.columns([2, 3, 2])
        with trade_c1:
            team_pick = st.selectbox("거래할 조", team_labels, key="trade_team")
        team_no = team_labels.index(team_pick) + 1
        with trade_c2:
            comp_options = [f"{c}  ·  {st.session_state.price_history[c][-1]:,}원" for c in company_pool]
            comp_pick = st.selectbox("종목 (현재가)", comp_options, key="trade_comp")
        comp_name = comp_pick.split("  ·  ")[0]
        cur_price = st.session_state.price_history[comp_name][-1]
        with trade_c3:
            trade_qty = st.number_input("수량 (주)", min_value=1, value=10, step=1, key="trade_qty")

        trade_amount = cur_price * trade_qty
        _pf = get_portfolio(team_no)
        _own_shares = _pf["holdings"].get(comp_name, {}).get("shares", 0)
        st.caption(
            f"주문 금액 **{trade_amount:,}원** · {team_no}조 보유 현금 **{_pf['cash']:,}원** · "
            f"{comp_name} 보유 **{_own_shares:,}주**"
        )

        buy_col, sell_col = st.columns(2)
        with buy_col:
            if st.button(f"🛒 매수 — {trade_amount:,}원", use_container_width=True, type="primary", key="btn_buy"):
                if trade_amount > _pf["cash"]:
                    st.error(f"현금이 부족합니다! (부족액 {trade_amount - _pf['cash']:,}원)")
                else:
                    h = _pf["holdings"].setdefault(comp_name, {"shares": 0, "avg_price": 0.0})
                    h["avg_price"] = (h["shares"] * h["avg_price"] + trade_qty * cur_price) / (h["shares"] + trade_qty)
                    h["shares"] += trade_qty
                    _pf["cash"] -= trade_amount
                    _pf["tx"].append({
                        "시점": f"{display_year}년 {display_q}분기", "조": f"{team_no}조", "구분": "매수",
                        "종목": comp_name, "수량": trade_qty, "단가(원)": cur_price, "금액(원)": trade_amount, "실현손익(원)": "",
                    })
                    save_current_state_to_file()
                    st.session_state.trade_msg = f"✅ {team_no}조가 {comp_name} {trade_qty:,}주를 {trade_amount:,}원에 매수했습니다."
                    st.rerun()
        with sell_col:
            if st.button(f"💸 매도 — {trade_amount:,}원", use_container_width=True, key="btn_sell"):
                if trade_qty > _own_shares:
                    st.error(f"보유 수량이 부족합니다! ({comp_name} 보유 {_own_shares:,}주)")
                else:
                    h = _pf["holdings"][comp_name]
                    realized = round((cur_price - h["avg_price"]) * trade_qty)
                    h["shares"] -= trade_qty
                    if h["shares"] == 0:
                        del _pf["holdings"][comp_name]
                    _pf["cash"] += trade_amount
                    _pf["tx"].append({
                        "시점": f"{display_year}년 {display_q}분기", "조": f"{team_no}조", "구분": "매도",
                        "종목": comp_name, "수량": trade_qty, "단가(원)": cur_price, "금액(원)": trade_amount, "실현손익(원)": realized,
                    })
                    save_current_state_to_file()
                    st.session_state.trade_msg = (
                        f"✅ {team_no}조가 {comp_name} {trade_qty:,}주를 매도했습니다. (실현손익 {realized:+,}원)"
                    )
                    st.rerun()

        # 📂 조별 포트폴리오 현황: 4개 조가 각각 무엇을 몇 주 들고 있는지 한눈에
        st.markdown("<div class='sec-title'>📂 조별 포트폴리오 현황</div>", unsafe_allow_html=True)
        pf_cols = st.columns(2)
        for idx, comp in enumerate(st.session_state.students_companies):
            no = idx + 1
            s = portfolio_summary(no)
            total_shares = sum(r["보유 수량"] for r in s["rows"])
            with pf_cols[idx % 2]:
                with st.container(border=True):
                    st.markdown(
                        f"**{no}조 · {comp}** &nbsp;|&nbsp; 🏦 총자산 **{s['total']:,}원** "
                        f"<span style='color:{'#2F9E44' if s['profit'] >= 0 else '#E03131'};'>"
                        f"({s['profit']:+,}원 · {s['return_pct']:+.2f}%)</span>",
                        unsafe_allow_html=True
                    )
                    st.caption(f"보유 {len(s['rows'])}종목 · 총 {total_shares:,}주 · 💵 현금 {s['cash']:,}원 · 📦 주식 {s['stock_value']:,}원")
                    if s["rows"]:
                        st.dataframe(pd.DataFrame(s["rows"]), use_container_width=True, hide_index=True)
                    else:
                        st.caption("아직 산 주식이 없어요 (전액 현금)")

        _all_tx = [t for no in range(1, len(st.session_state.students_companies) + 1) for t in get_portfolio(no)["tx"]]
        if _all_tx:
            with st.expander(f"🧾 전체 거래 내역 ({len(_all_tx)}건)"):
                tx_df = pd.DataFrame(_all_tx)
                st.dataframe(tx_df, use_container_width=True, hide_index=True)
                st.download_button(
                    "📥 거래 내역 CSV 다운로드", tx_df.to_csv(index=False).encode("utf-8-sig"),
                    file_name="조별_주식거래내역.csv", mime="text/csv",
                    key=f"dl_tx_{len(_all_tx)}"
                )
        st.stop()  # 주식 거래 탭에서는 아래 마케팅 평가 화면을 그리지 않음

    # ==========================================
    # [교사 탭 2] 📝 마케팅 전략 평가 (분기 진행)
    # ==========================================
    if st.session_state.quarter_records:
        with st.expander(f"📜 지난 분기 평가 기록 보기 (총 {len(st.session_state.quarter_records)}개 분기 진행됨)"):
            hist_rows = []
            for rec in st.session_state.quarter_records:
                for comp, r in rec["results"].items():
                    hist_rows.append({
                        "분기": rec["label"],
                        "시장 이슈": rec["issue"],
                        "기업": comp,
                        "AI 배수": r["multiplier"],
                        "제출 전략": r["strategy"],
                        "AI 코멘트": r["reason"],
                    })
            hist_df = pd.DataFrame(hist_rows)
            st.dataframe(hist_df, use_container_width=True, hide_index=True)
            st.download_button(
                "📥 평가 기록 CSV 다운로드 (수업 기록 보관용)",
                hist_df.to_csv(index=False).encode("utf-8-sig"),
                file_name="분기별_전략평가기록.csv", mime="text/csv",
                key=f"dl_records_{st.session_state.current_quarter}"
            )

    st.markdown("<div class='sec-title'>📝 조별 마케팅 기획안 입력</div>", unsafe_allow_html=True)
    strategies = {}
    for company in st.session_state.students_companies:
        strategies[company] = st.text_area(f"📝 {company} 마케팅 기획안", placeholder="조별로 제출한 마케팅 전략을 입력하세요...", height=100)

    # 빈 전략은 AI 평가 없이 ×1.0으로 처리되므로, 실수로 빈 채 진행하지 않도록 미리 알려줌
    _empty_teams = [c for c, s in strategies.items() if not s.strip()]
    if _empty_teams:
        st.warning(f"✏️ 전략이 아직 비어 있어요: **{', '.join(_empty_teams)}** — 빈 칸으로 진행하면 해당 조는 AI 평가 없이 주가 변동 없음(×1.0)으로 처리됩니다.")

    if st.button("🚀 AI 분석 실행 및 다음 분기 진행", type="primary", use_container_width=True):
        with st.spinner('AI가 거시 경제와 기업의 "현재 재무 지표"를 바탕으로 학생들의 기획안을 분석 중입니다...'):
            new_event = random.choice(market_events)
            st.session_state.market_issue = new_event
            event_text = new_event["이슈"]
            record_results = {}

            # AI 평가는 네트워크 대기가 대부분이므로 4개 조를 병렬 호출 (순차 대비 최대 4배 빠름).
            # 세션 상태를 읽는 준비 작업은 메인 스레드에서 먼저 끝내고, API 호출만 스레드로 보낸다.
            eval_jobs = [
                (company, company_pool[company]['industry'], strategies[company], get_financial_text_for_ai(company))
                for company in st.session_state.students_companies
            ]
            with ThreadPoolExecutor(max_workers=len(eval_jobs)) as pool:
                futures = {
                    company: pool.submit(evaluate_strategy_with_ai, company, industry, strategy, event_text, fin_text)
                    for company, industry, strategy, fin_text in eval_jobs
                }

            # 상태 업데이트(주가·실적 생성)는 순서 보장을 위해 메인 스레드에서 순차 처리
            for company, industry, strategy, fin_text in eval_jobs:
                multiplier, reason = futures[company].result()
                st.session_state.ai_evaluations[company] = {"score": multiplier, "reason": reason}
                record_results[company] = {"strategy": strategy, "multiplier": multiplier, "reason": reason}

                current_price = st.session_state.price_history[company][-1]
                st.session_state.price_history[company].append(int(current_price * multiplier))

                # 마케팅 성과가 반영된 이번 분기 '월별 실적 보고서' 생성 (직전 실적 기준으로 연속 변화)
                new_report = generate_quarter_report(company, st.session_state.current_quarter, multiplier)
                st.session_state.generated_reports.setdefault(company, []).extend(new_report)

            background_companies = [c for c in company_pool.keys() if c not in st.session_state.students_companies]
            for bg_comp in background_companies:
                bg_industry = company_pool[bg_comp]['industry']
                current_price = st.session_state.price_history[bg_comp][-1]
                base_impact = new_event["impact"].get(bg_industry, 1.0)
                noise = random.uniform(-0.05, 0.05)
                final_multiplier = max(0.6, min(1.4, base_impact + noise))
                st.session_state.price_history[bg_comp].append(int(current_price * final_multiplier))

            # 이번 분기 진행 내역을 기록에 누적 (시뮬레이션 결과 CSV의 원천 데이터)
            st.session_state.quarter_records.append({
                "quarter": st.session_state.current_quarter,
                "label": f"{display_year}년 {display_q}분기",
                "issue": event_text,
                "results": record_results
            })

            st.session_state.current_quarter += 1
            
            # 🌟 [중요] 모든 변동 사항 계산이 끝난 후 파일에 실시간 영구 백업
            save_current_state_to_file()
            
        st.balloons()
        st.success("데이터가 업데이트되었습니다. 학생 화면으로 이동해 결과를 확인하세요!")